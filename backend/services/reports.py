"""
Excel report generation and email to owner.
Generates a comprehensive Excel file with all service request data.
"""

import io
import logging
from datetime import datetime, timedelta

from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import ServiceRequest

logger = logging.getLogger(__name__)


def generate_excel_report():
    """
    Generate an Excel report with all service request data.
    Returns the workbook as bytes.
    """
    wb = Workbook()

    # ── Sheet 1: All Requests ──
    ws = wb.active
    ws.title = "All Requests"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Headers
    headers = [
        "ID", "Customer Name", "Phone", "Email", "Category",
        "Status", "Work Done?", "Assigned Expert", "Work Notes",
        "Conversation Step", "Created At", "Updated At", "User Account"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    requests = ServiceRequest.objects.select_related('category', 'user').all()
    for row_num, req in enumerate(requests, 2):
        row_data = [
            req.id,
            req.customer_name,
            req.customer_phone,
            req.customer_email,
            req.category.name,
            req.get_status_display(),
            "Yes ✓" if req.work_done else "No ✗",
            req.assigned_expert or "—",
            req.work_notes or "—",
            req.conversation_step,
            req.created_at.strftime("%Y-%m-%d %H:%M"),
            req.updated_at.strftime("%Y-%m-%d %H:%M"),
            req.user.email if req.user else "Anonymous",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

            # Color code status
            if col == 6:  # Status column
                if "Completed" in str(value):
                    cell.fill = PatternFill(start_color="D4EDDA", fill_type="solid")
                elif "Pending" in str(value):
                    cell.fill = PatternFill(start_color="FFF3CD", fill_type="solid")
                elif "Cancelled" in str(value):
                    cell.fill = PatternFill(start_color="F8D7DA", fill_type="solid")
            
            # Color code work done
            if col == 7:
                if req.work_done:
                    cell.fill = PatternFill(start_color="D4EDDA", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFF3CD", fill_type="solid")

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 35)

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Lagech Service Report Summary").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M IST')}")

    total = requests.count()
    completed = requests.filter(work_done=True).count()
    pending = requests.filter(status='pending').count()
    in_progress = requests.filter(status__in=['message_sent', 'in_conversation', 'assigned', 'in_progress']).count()
    cancelled = requests.filter(status='cancelled').count()

    summary_data = [
        ("", ""),
        ("Total Requests", total),
        ("Completed (Work Done)", completed),
        ("Pending", pending),
        ("In Progress", in_progress),
        ("Cancelled", cancelled),
        ("Completion Rate", f"{(completed / total * 100):.1f}%" if total > 0 else "N/A"),
    ]
    for row_num, (label, value) in enumerate(summary_data, 4):
        ws2.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row_num, column=2, value=value)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def email_report_to_owner():
    """
    Generate Excel report and email it to the owner.
    Called every hour by the scheduler.
    """
    logger.info("[Report] Generating hourly Excel report...")

    try:
        excel_bytes = generate_excel_report()
        now = timezone.now().strftime("%Y-%m-%d_%H-%M")

        email = EmailMessage(
            subject=f"Lagech Service Report – {now}",
            body=(
                f"Hi,\n\n"
                f"Please find attached the hourly service report for Lagech.\n"
                f"Generated at: {timezone.now().strftime('%Y-%m-%d %H:%M IST')}\n\n"
                f"This report contains:\n"
                f"• All service requests with their current status\n"
                f"• Work completion status\n"
                f"• Assigned experts\n"
                f"• Summary statistics\n\n"
                f"– Lagech System"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[settings.OWNER_EMAIL],
        )
        email.attach(
            f"lagech_report_{now}.xlsx",
            excel_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        email.send()
        logger.info(f"[Report] Sent to {settings.OWNER_EMAIL}")

    except Exception as e:
        logger.error(f"[Report] Failed to send: {e}")
