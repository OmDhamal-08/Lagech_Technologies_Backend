"""
Admin API views for Lagech dashboard.
Provides dashboard stats, orders list, and Excel export — restricted to admin email.
"""
import io
from datetime import timedelta
from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import ServiceRequest, Category


def is_admin(user):
    """Check if user is the admin."""
    return user.email.lower() == getattr(settings, 'ADMIN_EMAIL', '').lower()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_dashboard(request):
    """
    GET /api/admin/dashboard/
    Returns overview stats, orders by category, orders by status, recent orders.
    """
    if not is_admin(request.user):
        return Response({'error': 'Access denied'}, status=403)

    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())

    total = ServiceRequest.objects.count()
    today = ServiceRequest.objects.filter(created_at__gte=today_start).count()
    this_week = ServiceRequest.objects.filter(created_at__gte=week_start).count()
    completed = ServiceRequest.objects.filter(work_done=True).count()

    # Orders per category
    by_category = list(
        ServiceRequest.objects.values('category__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Orders per status
    by_status = list(
        ServiceRequest.objects.values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Recent 10 orders
    recent = ServiceRequest.objects.select_related('category').order_by('-created_at')[:10]
    recent_data = [
        {
            'id': r.id,
            'customer_name': r.customer_name or '—',
            'customer_phone': r.customer_phone,
            'category': r.category.name,
            'issue': r.message,
            'status': r.status,
            'status_display': r.get_status_display(),
            'date': r.created_at.strftime('%d %b %Y'),
            'time': r.created_at.strftime('%I:%M %p'),
            'created_at': r.created_at.isoformat(),
        }
        for r in recent
    ]

    return Response({
        'stats': {
            'total': total,
            'today': today,
            'this_week': this_week,
            'completed': completed,
        },
        'by_category': [
            {'name': c['category__name'], 'count': c['count']}
            for c in by_category
        ],
        'by_status': [
            {'status': s['status'], 'count': s['count']}
            for s in by_status
        ],
        'recent_orders': recent_data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_orders(request):
    """
    GET /api/admin/orders/
    Returns all orders with optional filters: category, status, date_from, date_to.
    """
    if not is_admin(request.user):
        return Response({'error': 'Access denied'}, status=403)

    qs = ServiceRequest.objects.select_related('category').order_by('-created_at')

    # Filters
    category = request.query_params.get('category')
    status_filter = request.query_params.get('status')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')

    if category:
        qs = qs.filter(category__name__iexact=category)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    data = [
        {
            'id': r.id,
            'customer_name': r.customer_name or '—',
            'customer_phone': r.customer_phone,
            'customer_email': r.customer_email,
            'category': r.category.name,
            'issue': r.message,
            'status': r.status,
            'status_display': r.get_status_display(),
            'notes': r.notes,
            'date': r.created_at.strftime('%d %b %Y'),
            'time': r.created_at.strftime('%I:%M %p'),
            'created_at': r.created_at.isoformat(),
        }
        for r in qs
    ]

    return Response({'orders': data, 'total': len(data)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_export_excel(request):
    """
    GET /api/admin/orders/export/
    Returns an Excel file (.xlsx) with all order data. Supports same filters as admin_orders.
    """
    if not is_admin(request.user):
        return Response({'error': 'Access denied'}, status=403)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    qs = ServiceRequest.objects.select_related('category').order_by('-created_at')

    # Same filters
    category = request.query_params.get('category')
    status_filter = request.query_params.get('status')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')

    if category:
        qs = qs.filter(category__name__iexact=category)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lagech Orders'

    # Header style
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['Order #', 'Customer Name', 'Phone', 'Email', 'Category', 'Issue', 'Status', 'Date', 'Time', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for row_idx, r in enumerate(qs, 2):
        row_data = [
            r.id,
            r.customer_name or '—',
            r.customer_phone,
            r.customer_email or '—',
            r.category.name,
            r.message,
            r.get_status_display(),
            r.created_at.strftime('%d %b %Y'),
            r.created_at.strftime('%I:%M %p'),
            r.notes,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Write to response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'lagech_orders_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
